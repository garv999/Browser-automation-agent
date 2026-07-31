"""Import the supplied *Test Orders* sheet into the agent's workbook schema.

The sheet the agent is given in real life does not look like the schema the agent
reads. Two differences matter, and both are structural rather than cosmetic:

1. **Different column names.** It has `Order Id`, `Return Window`, `Refund ID`,
   `No of Product`. Header matching here is case-, space- and punctuation-
   insensitive and alias-driven, so the file does not have to be edited by hand
   before a run.

2. **Multi-item orders live in one cell.** Order `OD337974610559` has four product
   links stacked in a single `Product Link` cell, interleaved with pasted chat
   text (`[8:23 pm, 26/06/2026] Arti Faym C: Take a look at this…`). One cell is
   four line items, and the whole point of the brief is that each gets its own
   row, its own return window and its own outcome. This splitter pulls every URL
   out of the cell, keys them by Flipkart's `pid`, and emits one row per SKU.

What it deliberately does **not** do is guess. Where the extracted SKU count
disagrees with the sheet's own `No of Product`, it warns and keeps every distinct
SKU it found rather than silently truncating to make the numbers line up.

Usage:
    python scripts/import_sheet.py "Test Orders.xlsx" --out data/return_tasks.xlsx
    python scripts/import_sheet.py "Test Orders.xlsx" --sheet Sheet1 --dry-run
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import load_workbook  # noqa: E402

from agent.excel_io import _as_date, _as_float, _as_int  # noqa: E402
from scripts.seed_workbook import write  # noqa: E402

URL_RE = re.compile(r"https?://[^\s\"'<>]+")
PID_RE = re.compile(r"[?&]pid=([A-Za-z0-9]+)", re.IGNORECASE)
#: Flipkart slugs sit between the host and `/p/`: .../dolsia-regular-women-blue-jeans/p/itm…
SLUG_RE = re.compile(r"https?://[^/]+/(?:dl/)?([a-z0-9-]{6,})/p/", re.IGNORECASE)

#: canonical column -> the header spellings seen in the wild
ALIASES = {
    "platform": ["platform"],
    "order_id": ["orderid", "orderno", "ordernumber"],
    "product_link": ["productlink", "producturl", "link", "product"],
    "amount": ["amount", "ordervalue", "totalamount", "price"],
    "quantity": ["noofproduct", "noofproducts", "qty", "quantity", "units"],
    "order_date": ["orderdate", "dateoforder"],
    "delivery_date": ["deliverydate", "delivereddate", "dateofdelivery"],
    "return_window_days": ["returnwindow", "returnwindowdays", "returnperiod"],
    "task_status": ["status", "taskstatus"],
    "return_id": ["refundid", "returnid", "rmaid"],
    "return_status": ["returnstatus"],
    "refund_amount": ["refundamount"],
    "timestamp": ["timestamp", "time"],
    "log": ["log", "notes", "remarks"],
    "address": ["address"],
    "contact_number": ["contactnumber", "phone", "mobile", "contact"],
}


def normalise(header: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(header or "").lower())


def map_headers(row) -> dict[str, int]:
    """Map canonical field -> column index, by alias."""
    found: dict[str, int] = {}
    for column, cell in enumerate(row, start=1):
        key = normalise(cell.value)
        if not key:
            continue
        for canonical, aliases in ALIASES.items():
            if key in aliases and canonical not in found:
                found[canonical] = column
    return found


def split_products(cell_value: str) -> list[dict[str, str]]:
    """Pull one entry per distinct product out of a possibly-messy cell.

    Deduplicated by `pid`, because the same product is often pasted twice (once
    as a shared chat link and once as a plain URL) and two rows for one SKU would
    mean attempting the same return twice.
    """
    text = str(cell_value or "")
    products: list[dict[str, str]] = []
    seen: set[str] = set()

    for url in URL_RE.findall(text):
        url = url.rstrip(".,);")
        pid_match = PID_RE.search(url)
        pid = pid_match.group(1) if pid_match else None
        key = pid or url
        if key in seen:
            continue
        seen.add(key)

        slug_match = SLUG_RE.search(url)
        name = slug_match.group(1).replace("-", " ").title() if slug_match else ""
        products.append({"link": url, "pid": pid or "", "name": name})

    if not products and text.strip():
        # A cell with no URL at all is still a line item — the agent will have to
        # match it by name, and flag it if it cannot.
        products.append({"link": "", "pid": "", "name": text.strip()[:120]})
    return products


def import_rows(path: Path, sheet_name: str | None, warn) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row, columns = None, {}
    # The real sheet does not always start on row 1, so find the header instead
    # of assuming it.
    for candidate in range(1, min(ws.max_row, 20) + 1):
        mapped = map_headers(ws[candidate])
        if "order_id" in mapped and "product_link" in mapped:
            header_row, columns = candidate, mapped
            break

    if header_row is None:
        raise SystemExit(
            "could not find a header row with recognisable 'Order Id' and 'Product Link' columns.\n"
            f"columns seen: {[c.value for c in ws[1]]}"
        )

    missing = [f for f in ("platform", "order_id", "product_link") if f not in columns]
    if missing:
        raise SystemExit(f"required column(s) not found: {', '.join(missing)}")

    rows: list[dict] = []
    for row in range(header_row + 1, ws.max_row + 1):

        def cell(field: str):
            index = columns.get(field)
            return ws.cell(row=row, column=index).value if index else None

        order_id = cell("order_id")
        if order_id in (None, ""):
            continue

        products = split_products(cell("product_link"))
        expected = _as_int(cell("quantity"), default=len(products))
        if expected != len(products):
            warn(
                f"order {order_id}: sheet says {expected} product(s), "
                f"{len(products)} distinct link(s) found — keeping all {len(products)}, "
                f"review before running"
            )

        total = _as_float(cell("amount"))
        per_item = round(total / len(products), 2) if total and products else total
        order_date = _as_date(cell("order_date"))
        delivery = _as_date(
            cell("delivery_date"), default_year=order_date.year if order_date else None
        )
        platform = str(cell("platform") or "Flipkart").strip().title()

        for product in products:
            rows.append(
                {
                    "platform": platform,
                    "order_id": str(order_id).strip(),
                    "sku": product["pid"],
                    "product_name": product["name"],
                    "product_link": product["link"],
                    "quantity": 1,
                    "amount": per_item,
                    "order_date": order_date,
                    "delivery_date": delivery,
                    "return_window_days": (
                        _as_int(cell("return_window_days"), default=None)
                        if cell("return_window_days") not in (None, "")
                        else None
                    ),
                    "address": cell("address"),
                    "contact_number": cell("contact_number"),
                }
            )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", help="the supplied Test Orders .xlsx")
    parser.add_argument("--sheet", default=None, help="worksheet name (default: the first)")
    parser.add_argument("--out", default=str(REPO_ROOT / "data" / "imported_tasks.xlsx"))
    parser.add_argument("--dry-run", action="store_true", help="report only, write nothing")
    args = parser.parse_args()

    warnings: list[str] = []

    def warn(message: str) -> None:
        warnings.append(message)
        print(f"  ! {message}")

    rows = import_rows(Path(args.source), args.sheet, warn)
    orders = len({r["order_id"] for r in rows})
    multi = sum(1 for oid in {r["order_id"] for r in rows} if sum(1 for r in rows if r["order_id"] == oid) > 1)

    print(f"\n{len(rows)} line item(s) across {orders} order(s); {multi} multi-item order(s)")
    no_window = sum(1 for r in rows if r["return_window_days"] is None)
    no_delivery = sum(1 for r in rows if r["delivery_date"] is None)
    if no_window:
        print(f"  ! {no_window} row(s) have no return window — eligibility will defer to the platform")
    if no_delivery:
        print(f"  ! {no_delivery} row(s) have no delivery date — these will go to human review")

    if args.dry_run:
        print("\ndry run: nothing written")
        return

    out = Path(args.out)
    write(rows, out)
    print(f"\nwrote {out}")


if __name__ == "__main__":
    main()
