"""Build the returns workbook.

Two profiles:

``--source mock``
    Derives rows from `mock_site/fixtures.py`, so the sheet and the mock
    storefront agree on every order, SKU, delivery date and return window. This
    is the workbook the test suite drives.

``--source sheet``
    A transcription of the supplied *Test Orders* sheet, for running against the
    live platforms. The interesting part is what happens to its multi-item rows:
    the source sheet crams four product links into a single cell, and this
    seeder explodes them into one row per SKU — which is the whole point of the
    per-line-item requirement. Where the source listed a link annotated "NA",
    that link is dropped and the row count follows the sheet's "No of Product".

Usage:
    python scripts/seed_workbook.py --source mock  --out data/return_tasks.xlsx
    python scripts/seed_workbook.py --source sheet --out data/test_orders_from_sheet.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, timedelta
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))  # so the script runs from anywhere

from openpyxl import load_workbook  # noqa: E402

from agent.excel_io import COLUMNS, create_workbook  # noqa: E402
from mock_site.fixtures import ALL_ORDERS  # noqa: E402

FLIPKART_LINK = "https://www.flipkart.com/p/itm{slug}?pid={pid}&marketplace=FLIPKART"
AMAZON_LINK = "https://www.amazon.in/dp/{pid}"

# Transcribed verbatim from the supplied Test Orders sheet, with its own dates
# rather than dates synthesised relative to the seeding run. Order IDs are the
# full 20-character Flipkart values (OD + 18 digits); the sheet wraps them across
# two lines in the Order Id column, and a transcription that stops at the wrap
# produces an ID the platform cannot resolve.
#
# Where the sheet gives a delivery *range* ("5-6 July") the later day is used, on
# the same reasoning as the parser: assuming the earlier day shortens the window
# and risks skipping an item that is still returnable.
SHEET_ROWS = [
    # (order_id, pid, product_name, amount, qty, window_days, order_date, delivery_date)
    ("OD337915012166989100", "TSHG9FQZSSAUGKUP", "Wearza Colorblock Men Round Neck Black Yellow T-shirt", 284, 1, 10, date(2026, 6, 24), date(2026, 6, 27)),
    ("OD337915105120141100", "VPAHNMQYW9PYYWH8", "Gulab Thar Women Kurta", 350, 1, 7, date(2026, 6, 24), date(2026, 6, 27)),
    ("OD337915117555423100", "KTAHKYRPHHMY4FK3", "Vedika Fashiondesgin Women Striped Printed Flared Kurta", 355, 1, 10, date(2026, 6, 24), date(2026, 6, 27)),
    ("OD337960018546978100", "ETHH7Z3FJTCRQQNB", "Bunaai Dhaga Women Kurta Palazzo Dupatta Set", 686, 1, 10, date(2026, 6, 29), date(2026, 7, 2)),
    # One order, four SKUs — the sheet held these in a single cell. Amounts split
    # from the order total of 2579 across its 4 products.
    ("OD337974610559997100", "JEAHJHY3CBJYNZNW", "Dolsia Regular Women Blue Jeans", 645, 1, 10, date(2026, 7, 1), date(2026, 7, 6)),
    ("OD337974610559997100", "JEAH87B2GRCCS3DZ", "Tokyo Talkies Loose Fit Women Blue Jeans", 645, 1, 10, date(2026, 7, 1), date(2026, 7, 6)),
    ("OD337974610559997100", "DREHHF5SKMVFGUKU", "Vasan Women A-Line Maroon Midi Calf Length Dress", 645, 1, 10, date(2026, 7, 1), date(2026, 7, 6)),
    ("OD337974610559997100", "DREHK6H2PN8XX6ZM", "Shivanshcloset Women Fit Flare Blue Beige Maxi Full Length Dress", 644, 1, 10, date(2026, 7, 1), date(2026, 7, 6)),
    # Two sarees on one order; total 1909.
    ("OD337983106511516100", "SARHMZB7GBADA4FK", "Arti Faym Embroidered Bollywood Satin Silk Blend Saree", 955, 1, 10, date(2026, 7, 2), date(2026, 7, 7)),
    ("OD337983106511516100", "SARHRKVNUD2PSCGY", "Brahmani Creation Embroidered Bollywood Georgette Saree", 954, 1, 10, date(2026, 7, 2), date(2026, 7, 7)),
    # Four shoulder bags; total 1582. Not yet delivered when the sheet was taken.
    ("OD337983703007211100", "HMBH8MV7VA3PCJDQ", "Carrylux Women Pink Shoulder Bag", 395, 1, 10, date(2026, 7, 2), date(2026, 7, 9)),
    ("OD337983703007211100", "HMBH8KTQFP5BPSZG", "Carrylux Women Beige Shoulder Bag", 395, 1, 10, date(2026, 7, 2), date(2026, 7, 9)),
    ("OD337983703007211100", "HMBH7MYANGQFUKJH", "Carrylux Women Red Shoulder Bag", 396, 1, 10, date(2026, 7, 2), date(2026, 7, 9)),
    ("OD337983703007211100", "HMBH8BNZ6JXFJCD9", "Carrylux Women Black Shoulder Bag", 396, 1, 10, date(2026, 7, 2), date(2026, 7, 9)),
]

SHEET_ADDRESS = "Sample Address, Sector 42, Gurgaon, Haryana - 122009, India"
SHEET_CONTACT = "9000000001"


def _slug(name: str) -> str:
    return "".join(ch.lower() for ch in name if ch.isalnum())[:24]


def rows_from_fixtures() -> list[dict]:
    rows: list[dict] = []
    for order in ALL_ORDERS:
        for item in order.items:
            link = (
                AMAZON_LINK.format(pid=item.product_id)
                if order.platform == "amazon"
                else FLIPKART_LINK.format(slug=_slug(item.title), pid=item.product_id)
            )
            rows.append(
                {
                    "platform": "Amazon" if order.platform == "amazon" else "Flipkart",
                    "order_id": order.order_id,
                    "sku": item.product_id,
                    "product_name": item.title,
                    "product_link": link,
                    "quantity": item.quantity,
                    "amount": item.amount * item.quantity,
                    "order_date": item.order_date,
                    "delivery_date": item.delivery_date,
                    "return_window_days": item.return_window_days,
                    "address": order.address,
                    "contact_number": order.contact,
                }
            )
    return rows


def rows_from_sheet() -> list[dict]:
    """The sheet's own orders, with the sheet's own dates.

    These are historical: the sheet was captured in early July 2026, so most of
    its return windows have since closed. That is a true statement about the
    data, and the agent reporting "out of window" for them is the correct result
    rather than a bug in the seeder.
    """
    rows: list[dict] = []
    for order_id, pid, name, amount, qty, window, ordered, delivered in SHEET_ROWS:
        rows.append(
            {
                "platform": "Flipkart",
                "order_id": order_id,
                "sku": pid,
                "product_name": name,
                "product_link": FLIPKART_LINK.format(slug=_slug(name), pid=pid),
                "quantity": qty,
                "amount": amount * qty,
                "order_date": ordered,
                "delivery_date": delivered,
                "return_window_days": window,
                "address": SHEET_ADDRESS,
                "contact_number": SHEET_CONTACT,
            }
        )
    return rows


def write(rows: list[dict], out: Path, sheet_name: str = "Returns") -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.exists():
        out.unlink()
    create_workbook(out, sheet_name)

    wb = load_workbook(out)
    ws = wb[sheet_name]
    index = {name: i + 1 for i, name in enumerate(COLUMNS)}

    for number, row in enumerate(rows, start=1):
        target = number + 1  # header occupies row 1
        ws.cell(row=target, column=index["task_id"]).value = f"T{number:03d}"
        ws.cell(row=target, column=index["task_status"]).value = "Pending"
        for key, value in row.items():
            ws.cell(row=target, column=index[key]).value = value

    for column, width in {
        "product_link": 60,
        "product_name": 46,
        "order_id": 20,
        "address": 40,
        "log": 60,
        "return_status": 24,
        "task_status": 20,
    }.items():
        ws.column_dimensions[ws.cell(row=1, column=index[column]).column_letter].width = width

    wb.save(out)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", choices=["mock", "sheet"], default="mock")
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    rows = rows_from_fixtures() if args.source == "mock" else rows_from_sheet()
    default = "return_tasks.xlsx" if args.source == "mock" else "test_orders_from_sheet.xlsx"
    out = Path(args.out) if args.out else REPO_ROOT / "data" / default

    write(rows, out)
    orders = len({r["order_id"] for r in rows})
    print(f"wrote {len(rows)} line item(s) across {orders} order(s) -> {out}")


if __name__ == "__main__":
    main()
