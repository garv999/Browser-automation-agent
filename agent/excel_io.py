"""Workbook read/write.

Two properties matter here and both are about not losing work:

1. **Write-back is per line item.** `write_outcome` touches exactly the row the
   task came from. An order never gets a single collective verdict.
2. **Write-back is immediate and durable.** The outcome is flushed to disk as
   soon as it is known, before the agent moves on. If the process dies halfway
   through an order, every return already placed on the platform is recorded in
   the sheet — the failure mode we must never have is a return that exists on
   Flipkart but not in the file.
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Platform, ReturnOutcome, ReturnStatus, ReturnTask, TaskStatus

#: Canonical column order. Input columns come first, agent-written columns last,
#: mirroring the write-back table in the spec.
COLUMNS = [
    "task_id",
    "platform",
    "order_id",
    "sku",
    "product_name",
    "product_link",
    "quantity",
    "amount",
    "order_date",
    "delivery_date",
    "return_window_days",
    "task_status",
    "return_id",
    "return_status",
    "refund_amount",
    "timestamp",
    "log",
    "address",
    "contact_number",
]

AGENT_WRITTEN = {"task_status", "return_id", "return_status", "refund_amount", "timestamp", "log"}

HEADER_ROW = 1


class WorkbookError(RuntimeError):
    pass


def create_workbook(path: Path, sheet_name: str = "Returns") -> None:
    """Create an empty workbook with the canonical header."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(COLUMNS)
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _header_map(ws: Worksheet) -> dict[str, int]:
    header = {}
    for col, cell in enumerate(ws[HEADER_ROW], start=1):
        if cell.value:
            header[str(cell.value).strip()] = col
    missing = [c for c in COLUMNS if c not in header]
    if missing:
        raise WorkbookError(f"workbook is missing column(s): {', '.join(missing)}")
    return header


#: "5-6 July" — a delivery *range*, which the supplied sheet uses freely.
DATE_RANGE_RE = re.compile(r"^\s*(\d{1,2})\s*[-–—to]+\s*(\d{1,2})\s+(\w+)", re.IGNORECASE)

DATE_FORMATS_WITH_YEAR = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %B %Y", "%d %b %Y", "%B %d %Y", "%b %d %Y")
DATE_FORMATS_NO_YEAR = ("%d %B", "%d %b", "%B %d", "%b %d")


def _as_date(value, default_year: Optional[int] = None) -> Optional[date]:
    """Parse a date the way the supplied sheet actually writes them.

    That sheet contains `24 June 2026`, `27 June` (no year at all) and
    `5-6 July` (a two-day delivery window). A parser that only accepts ISO dates
    returns None for three of those, and a None delivery date makes every row
    ineligible — the agent would refuse to do any work on the real file.

    For a range, the **later** day is taken. Assuming the earlier one would shrink
    the computed return window and risk skipping an item that is still returnable;
    assuming the later one at worst costs a page load, and the platform — which is
    the authority — refuses the item and that refusal gets recorded.
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()

    match = DATE_RANGE_RE.match(text)
    if match:
        text = f"{match.group(2)} {match.group(3)}"

    for fmt in DATE_FORMATS_WITH_YEAR:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    # No year on the value: take it from the order date where we have one, so a
    # sheet written in December about a January delivery does not land a year out.
    # The year is appended to the *text* rather than patched onto the result —
    # strptime without a year is deprecated in 3.15 and mishandles 29 February.
    year = default_year or date.today().year
    for fmt in DATE_FORMATS_NO_YEAR:
        try:
            return datetime.strptime(f"{text} {year}", f"{fmt} %Y").date()
        except ValueError:
            continue
    return None


def _as_datetime(value) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _as_float(value) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").replace("₹", "").strip())
    except ValueError:
        return None


def _as_int(value, default: int = 1) -> int:
    """First integer in the value.

    The supplied sheet writes return windows as `10 Days`, `7 Days`, `10 Day`.
    Requiring a bare number turns all of those into "no window on record", which
    silently disables the eligibility check for every row.
    """
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"-?\d+", str(value))
    return int(match.group()) if match else default


#: Task-status wordings treated as work still to do. The brief specifies
#: "To Do" / "Pending"; the rest are the common variants a hand-kept sheet uses.
PENDING_ALIASES = {"", "pending", "to do", "todo", "to-do", "open", "new", "not started", "none"}
DONE_ALIASES = {"done", "complete", "completed", "closed", "finished"}
REVIEW_ALIASES = {"needs human review", "needs review", "review", "manual", "human review", "escalate"}


def _as_task_status(value) -> TaskStatus:
    """Classify the sheet's task-status column.

    An unrecognised value is deliberately **not** treated as pending. Defaulting
    an unknown word to "pending" means a sheet whose finished rows say "Complete"
    would have every one of them attempted again, placing a second return on
    orders already returned. Since that is irreversible, anything this function
    does not recognise is routed to human review instead.
    """
    text = str(value).strip().lower() if value is not None else ""
    if text in PENDING_ALIASES:
        return TaskStatus.PENDING
    if text in DONE_ALIASES:
        return TaskStatus.DONE
    if text in REVIEW_ALIASES:
        return TaskStatus.NEEDS_REVIEW
    return TaskStatus.NEEDS_REVIEW


def _as_enum(enum_cls, value, default=None):
    if value in (None, ""):
        return default
    text = str(value).strip()
    for member in enum_cls:
        if member.value.lower() == text.lower():
            return member
    return default


def read_tasks(path: Path, sheet_name: str = "Returns") -> list[ReturnTask]:
    """Read every row of the sheet into `ReturnTask`s, pending or not."""
    path = Path(path)
    if not path.exists():
        raise WorkbookError(f"workbook not found: {path}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise WorkbookError(f"sheet {sheet_name!r} not in {wb.sheetnames}")
    ws = wb[sheet_name]
    header = _header_map(ws)

    tasks: list[ReturnTask] = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):

        def cell(name: str):
            return ws.cell(row=row, column=header[name]).value

        order_id = cell("order_id")
        if order_id in (None, ""):
            continue  # blank spacer row

        # Parsed first so a year-less delivery date ("27 June") can borrow its year.
        order_date = _as_date(cell("order_date"))

        platform = _as_enum(Platform, cell("platform"))
        if platform is None:
            raise WorkbookError(f"row {row}: unrecognised platform {cell('platform')!r}")

        tasks.append(
            ReturnTask(
                row=row,
                platform=platform,
                order_id=str(order_id).strip(),
                product_link=str(cell("product_link") or "").strip(),
                sku=(str(cell("sku")).strip() if cell("sku") else None),
                product_name=(str(cell("product_name")).strip() if cell("product_name") else None),
                quantity=_as_int(cell("quantity")),
                amount=_as_float(cell("amount")),
                order_date=order_date,
                delivery_date=_as_date(
                    cell("delivery_date"), default_year=order_date.year if order_date else None
                ),
                return_window_days=(
                    _as_int(cell("return_window_days"), default=None)
                    if cell("return_window_days") not in (None, "")
                    else None
                ),
                task_status=_as_task_status(cell("task_status")),
                return_id=(str(cell("return_id")).strip() if cell("return_id") else None),
                return_status=_as_enum(ReturnStatus, cell("return_status")),
                refund_amount=_as_float(cell("refund_amount")),
                timestamp=_as_datetime(cell("timestamp")),
                log=str(cell("log") or ""),
                notes={
                    "address": cell("address"),
                    "contact_number": cell("contact_number"),
                    "task_id": cell("task_id"),
                },
            )
        )
    return tasks


def pending_tasks(tasks: Iterable[ReturnTask]) -> list[ReturnTask]:
    return [t for t in tasks if t.is_pending]


def _atomic_save(wb: Workbook, path: Path) -> None:
    """Save via a temp file + replace so a crash mid-write cannot truncate the
    workbook and lose already-recorded returns."""
    fd, tmp = tempfile.mkstemp(dir=str(path.parent), suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        shutil.move(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)


def write_outcome(
    path: Path,
    task: ReturnTask,
    outcome: ReturnOutcome,
    sheet_name: str = "Returns",
    when: Optional[datetime] = None,
) -> None:
    """Write one line item's result back to its own row and flush to disk."""
    path = Path(path)
    when = when or datetime.now()

    wb = load_workbook(path)
    ws = wb[sheet_name]
    header = _header_map(ws)

    def put(name: str, value) -> None:
        ws.cell(row=task.row, column=header[name]).value = value

    put("return_id", outcome.return_id or "")
    put("return_status", outcome.return_status.value)
    put("refund_amount", outcome.refund_amount if outcome.refund_amount is not None else "")
    put("task_status", outcome.task_status.value)
    put("timestamp", when.strftime("%Y-%m-%d %H:%M:%S"))

    existing = str(ws.cell(row=task.row, column=header["log"]).value or "").strip()
    entry = f"[{when:%Y-%m-%d %H:%M:%S}] {outcome.message}".strip()
    put("log", f"{existing}\n{entry}".strip() if existing else entry)

    _atomic_save(wb, path)

    # Keep the in-memory task consistent with what is now on disk.
    task.return_id = outcome.return_id
    task.return_status = outcome.return_status
    task.refund_amount = outcome.refund_amount
    task.task_status = outcome.task_status
    task.timestamp = when
    task.log = str(ws.cell(row=task.row, column=header["log"]).value or "")


def order_is_fully_settled(tasks: Iterable[ReturnTask], order_id: str) -> bool:
    """An order is only 'done' once every one of its line items has a final
    recorded state — the spec's last partial-success rule."""
    items = [t for t in tasks if t.order_id == order_id]
    return bool(items) and all(t.task_status != TaskStatus.PENDING for t in items)
