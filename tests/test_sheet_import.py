"""Importing the supplied Test Orders sheet.

The fixture below is deliberately built to look like the real file rather than
like a tidy one: the sheet's own column names, return windows written as
"10 Days", delivery dates with no year and as two-day ranges, and a multi-item
order whose four product links share one cell with pasted chat text.
"""

from __future__ import annotations

from datetime import date

import pytest
from openpyxl import Workbook

from agent.excel_io import _as_date, _as_int, read_tasks
from scripts.import_sheet import import_rows, split_products

# One cell, four products, exactly as the supplied sheet stores them — including
# a duplicate paste of the first link and interleaved chat noise.
MESSY_CELL = """[8:23 pm, 26/06/2026] Arti Faym C: Take a look at this
https://www.flipkart.com/dolsia-regular-women-blue-jeans/p/itme00ed7c7bdab3?pid=JEAHJHY3CBJYNZNW&lid=LSTJEAHJHY3CBJYNZNW40T4CZ&marketplace=FLIPKART

https://www.flipkart.com/tokyo-talkies-loose-fit-women-blue-jeans/p/itmba1701124098b?pid=JEAH87B2GRCCS3DZ&marketplace=FLIPKART  NA

https://www.flipkart.com/dolsia-regular-women-blue-jeans/p/itme00ed7c7bdab3?pid=JEAHJHY3CBJYNZNW&lid=OTHER

https://www.flipkart.com/vasan-women-a-line-maroon-midi-calf-length-dress/p/itm9dd3d2688a45a?pid=DREHHF5SKMVFGUKU&marketplace=FLIPKART
[8:25 pm, 26/06/2026] Arti Faym C: and this one
https://www.flipkart.com/shivanshcloset-women-fit-flare-blue-beige-maxi-full-length-dress/p/itm30fb421b5f582?pid=DREHK6H2PN8XX6ZM&marketplace=FLIPKART
"""

HEADERS = [
    "Address", "Contact Number", "Product Link", "Amount", "No of Product",
    "Order date", "Order Id", "Delivery date", "Return Window", "Status",
    "Platform", "Refund ID", "Return Status", "Refund Amount", "Timestamp", "Log",
]


@pytest.fixture
def supplied_sheet(tmp_path):
    """A workbook shaped like the file the agent is actually handed."""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)

    ws.append([
        "Sample Address, Sector 42, Gurgaon", "9000000001",
        "https://www.flipkart.com/wearza-colorblock-men-round-neck-black-yellow-t-shirt/p/itm709?pid=TSHG9FQZSSAUGKUP&marketplace=FLIPKART",
        284, 1, "24 June 2026", "OD337915012166", "27 June", "10 Days", "Pending",
        "Flipkart", "", "", "", "", "",
    ])
    ws.append([
        "Sample Address, Sector 42, Gurgaon", "9000000001",
        "https://www.flipkart.com/gulab-thar/p/itm076?pid=VPAHNMQYW9PYYWH8&marketplace=FLIPKART",
        350, 1, "24 June 2026", "OD337915105120", "27 June", "7 Days", "Pending",
        "Flipkart", "", "", "", "", "",
    ])
    ws.append([
        "Sample Address 2, New Delhi - 110041", "9000000002",
        MESSY_CELL, 2579, 4, "01 July 2026", "OD337974610559", "5-6 July", "10 Days",
        "Pending", "Flipkart", "", "", "", "", "",
    ])

    path = tmp_path / "Test Orders.xlsx"
    wb.save(path)
    return path


# -- the value formats the real sheet uses --------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [("10 Days", 10), ("7 Days", 7), ("10 Day", 10), ("10 days", 10), ("10", 10), (10, 10), (10.0, 10)],
)
def test_return_window_parses_the_sheets_own_wording(raw, expected):
    """`10 Days` returning None silently disables the eligibility check."""
    assert _as_int(raw, default=None) == expected


def test_dateless_year_is_taken_from_the_order_date():
    assert _as_date("27 June", default_year=2026) == date(2026, 6, 27)


def test_delivery_range_takes_the_later_day():
    """`5-6 July` — assuming the earlier day would shrink the window and risk
    skipping an item that is still returnable."""
    assert _as_date("5-6 July", default_year=2026) == date(2026, 7, 6)
    assert _as_date("8-9 July", default_year=2026) == date(2026, 7, 9)


def test_real_dates_still_parse():
    assert _as_date("24 June 2026") == date(2026, 6, 24)
    assert _as_date("2026-06-24") == date(2026, 6, 24)
    assert _as_date(date(2026, 6, 24)) == date(2026, 6, 24)


def test_unparseable_values_are_still_none():
    assert _as_date("sometime next week") is None
    assert _as_date("") is None


# -- splitting a multi-item cell ------------------------------------------


def test_one_cell_becomes_one_entry_per_distinct_product():
    products = split_products(MESSY_CELL)
    assert len(products) == 4, "the duplicated link must not become a fifth item"
    assert [p["pid"] for p in products] == [
        "JEAHJHY3CBJYNZNW",
        "JEAH87B2GRCCS3DZ",
        "DREHHF5SKMVFGUKU",
        "DREHK6H2PN8XX6ZM",
    ]


def test_product_names_come_from_the_url_slug():
    products = split_products(MESSY_CELL)
    assert products[0]["name"] == "Dolsia Regular Women Blue Jeans"
    assert products[2]["name"] == "Vasan Women A Line Maroon Midi Calf Length Dress"


def test_chat_text_is_not_mistaken_for_a_product():
    for product in split_products(MESSY_CELL):
        assert product["link"].startswith("http")
        assert "Arti Faym" not in product["name"]


def test_a_cell_with_no_url_is_still_a_line_item():
    """Dropping it would silently lose a return the user is owed."""
    products = split_products("Blue kurta, size XXL — no link available")
    assert len(products) == 1
    assert products[0]["pid"] == ""
    assert "Blue kurta" in products[0]["name"]


# -- the importer end to end ----------------------------------------------


def test_supplied_sheet_imports_to_line_items(supplied_sheet, tmp_path):
    warnings: list[str] = []
    rows = import_rows(supplied_sheet, None, warnings.append)

    # 2 single-item orders + 4 SKUs on the multi-item order.
    assert len(rows) == 6
    multi = [r for r in rows if r["order_id"] == "OD337974610559"]
    assert len(multi) == 4
    assert len({r["sku"] for r in multi}) == 4


def test_imported_windows_and_dates_are_usable(supplied_sheet):
    rows = import_rows(supplied_sheet, None, lambda m: None)
    by_order = {r["order_id"]: r for r in rows}

    assert by_order["OD337915012166"]["return_window_days"] == 10
    assert by_order["OD337915012166"]["delivery_date"] == date(2026, 6, 27)
    assert by_order["OD337915105120"]["return_window_days"] == 7
    assert by_order["OD337974610559"]["delivery_date"] == date(2026, 7, 6)

    # Nothing may be left without the two fields eligibility depends on.
    assert all(r["delivery_date"] is not None for r in rows)
    assert all(r["return_window_days"] is not None for r in rows)


def test_order_total_is_split_across_its_line_items(supplied_sheet):
    rows = import_rows(supplied_sheet, None, lambda m: None)
    multi = [r for r in rows if r["order_id"] == "OD337974610559"]
    assert all(r["amount"] == pytest.approx(2579 / 4, abs=0.01) for r in multi)


def test_count_mismatch_warns_instead_of_truncating(tmp_path):
    """The sheet says 4 products but only 3 links survive deduplication. Keeping
    3 and saying so beats inventing a 4th or dropping one to make it fit."""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append([
        "addr", "9000000001",
        "https://f.com/a/p/i?pid=AAA\nhttps://f.com/b/p/i?pid=BBB\nhttps://f.com/c/p/i?pid=CCC",
        900, 4, "01 July 2026", "OD1", "5 July", "10 Days", "Pending",
        "Flipkart", "", "", "", "", "",
    ])
    path = tmp_path / "mismatch.xlsx"
    wb.save(path)

    warnings: list[str] = []
    rows = import_rows(path, None, warnings.append)

    assert len(rows) == 3
    assert any("sheet says 4 product(s), 3 distinct link(s)" in w for w in warnings)


def test_imported_workbook_is_readable_by_the_agent(supplied_sheet, tmp_path):
    """The whole point: the imported file must load straight into the runner."""
    from scripts.seed_workbook import write

    rows = import_rows(supplied_sheet, None, lambda m: None)
    out = tmp_path / "imported.xlsx"
    write(rows, out)

    tasks = read_tasks(out)
    assert len(tasks) == 6
    assert all(t.is_pending for t in tasks)
    assert all(t.return_window_days is not None for t in tasks)
    assert {t.sku for t in tasks if t.order_id == "OD337974610559"} == {
        "JEAHJHY3CBJYNZNW", "JEAH87B2GRCCS3DZ", "DREHHF5SKMVFGUKU", "DREHK6H2PN8XX6ZM",
    }


def test_headers_match_regardless_of_case_and_spacing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["ORDER ID", "product_link", "PLATFORM", "Return  Window", "Delivery Date", "Order Date"])
    ws.append(["OD9", "https://f.com/x/p/i?pid=ZZZ", "flipkart", "10 Days", "5 July", "1 July 2026"])
    path = tmp_path / "odd_headers.xlsx"
    wb.save(path)

    rows = import_rows(path, None, lambda m: None)
    assert len(rows) == 1
    assert rows[0]["sku"] == "ZZZ"
    assert rows[0]["platform"] == "Flipkart"
    assert rows[0]["return_window_days"] == 10
