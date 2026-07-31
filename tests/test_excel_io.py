"""Workbook read/write.

The property that matters: an outcome lands on exactly one row, and it is on
disk before the agent takes another action.
"""

from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import load_workbook

from agent import excel_io
from agent.excel_io import WorkbookError, create_workbook, read_tasks, write_outcome
from agent.models import Platform, ReturnOutcome, ReturnStatus, TaskStatus


def test_seeded_workbook_reads_back_as_line_items(workbook):
    tasks = read_tasks(workbook)
    assert tasks, "seeded workbook should not be empty"
    assert all(t.task_status == TaskStatus.PENDING for t in tasks)

    # The four-SKU Flipkart order must be four separate rows, not one.
    multi = [t for t in tasks if t.order_id == "OD337974610559"]
    assert len(multi) == 4
    assert len({t.row for t in multi}) == 4
    assert len({t.sku for t in multi}) == 4


def test_write_outcome_touches_only_its_own_row(workbook):
    tasks = read_tasks(workbook)
    target = next(t for t in tasks if t.order_id == "OD337974610559")
    siblings = [t for t in tasks if t.order_id == "OD337974610559" and t.row != target.row]

    write_outcome(
        workbook,
        target,
        ReturnOutcome(ReturnStatus.PLACED, "CR26060000000001", 645.0, "return placed"),
        when=datetime(2026, 7, 31, 12, 0, 0),
    )

    after = {t.row: t for t in read_tasks(workbook)}
    assert after[target.row].return_id == "CR26060000000001"
    assert after[target.row].task_status == TaskStatus.DONE
    assert after[target.row].refund_amount == 645.0

    for sibling in siblings:
        assert after[sibling.row].task_status == TaskStatus.PENDING
        assert after[sibling.row].return_id is None


def test_outcome_is_flushed_to_disk_immediately(workbook):
    """A crash after this call must not lose a return that already exists on
    the platform, so the write cannot be buffered in memory."""
    task = read_tasks(workbook)[0]
    write_outcome(workbook, task, ReturnOutcome(ReturnStatus.PLACED, "CR1234567890123", 100.0))

    raw = load_workbook(workbook)["Returns"]
    header = {c.value: c.column for c in raw[1]}
    assert raw.cell(row=task.row, column=header["return_id"]).value == "CR1234567890123"


def test_log_entries_accumulate_rather_than_overwrite(workbook):
    task = read_tasks(workbook)[0]
    write_outcome(workbook, task, ReturnOutcome(ReturnStatus.FAILED, message="first attempt failed"))
    write_outcome(workbook, task, ReturnOutcome(ReturnStatus.PLACED, "CR9999999999999", message="retry succeeded"))

    log = read_tasks(workbook)[0].log
    assert "first attempt failed" in log
    assert "retry succeeded" in log


def test_needs_review_statuses_map_to_needs_review():
    assert ReturnOutcome(ReturnStatus.FAILED).task_status == TaskStatus.NEEDS_REVIEW
    assert ReturnOutcome(ReturnStatus.SUPPORT_NEEDED).task_status == TaskStatus.NEEDS_REVIEW
    assert ReturnOutcome(ReturnStatus.NOT_DELIVERED).task_status == TaskStatus.NEEDS_REVIEW
    # Out of window is a correct, final answer — not something a human must fix.
    assert ReturnOutcome(ReturnStatus.OUT_OF_WINDOW).task_status == TaskStatus.DONE
    assert ReturnOutcome(ReturnStatus.PLACED).task_status == TaskStatus.DONE


def test_order_is_settled_only_when_every_item_has_a_state(workbook):
    tasks = read_tasks(workbook)
    order = "OD337974610559"
    items = [t for t in tasks if t.order_id == order]

    assert not excel_io.order_is_fully_settled(tasks, order)

    for item in items[:-1]:
        item.task_status = TaskStatus.DONE
    assert not excel_io.order_is_fully_settled(tasks, order)

    items[-1].task_status = TaskStatus.NEEDS_REVIEW
    assert excel_io.order_is_fully_settled(tasks, order)


@pytest.mark.parametrize("raw", ["Pending", "To Do", "todo", "TO-DO", "Open", "New", "", None])
def test_work_still_to_do_is_recognised_as_pending(raw):
    """The brief specifies status = "To Do" / "Pending"; a hand-kept sheet uses
    variants of both."""
    assert excel_io._as_task_status(raw) == TaskStatus.PENDING


@pytest.mark.parametrize("raw", ["Done", "Complete", "Completed", "Closed", "finished"])
def test_finished_work_is_never_re_attempted(raw):
    """The dangerous direction. If "Complete" fell through to pending, every
    finished row would be returned a second time, and a return is irreversible."""
    assert excel_io._as_task_status(raw) == TaskStatus.DONE


@pytest.mark.parametrize("raw", ["Blocked", "???", "on hold", "waiting for courier"])
def test_unrecognised_status_goes_to_a_human_not_to_the_platform(raw):
    assert excel_io._as_task_status(raw) == TaskStatus.NEEDS_REVIEW


def test_a_sheet_marked_complete_yields_no_pending_work(tmp_path):
    """End to end on the workbook, not just the helper."""
    path = tmp_path / "complete.xlsx"
    create_workbook(path)
    wb = load_workbook(path)
    ws = wb["Returns"]
    index = {name: i + 1 for i, name in enumerate(excel_io.COLUMNS)}
    for row, status in enumerate(["Complete", "Closed", "To Do"], start=2):
        ws.cell(row=row, column=index["order_id"]).value = f"OD{row}"
        ws.cell(row=row, column=index["platform"]).value = "Flipkart"
        ws.cell(row=row, column=index["task_status"]).value = status
    wb.save(path)

    pending = excel_io.pending_tasks(read_tasks(path))
    assert [t.order_id for t in pending] == ["OD4"]


def test_missing_column_is_rejected_loudly(tmp_path):
    path = tmp_path / "broken.xlsx"
    create_workbook(path)
    wb = load_workbook(path)
    wb["Returns"].cell(row=1, column=1).value = "not_task_id"
    wb.save(path)

    with pytest.raises(WorkbookError, match="missing column"):
        read_tasks(path)


def test_unknown_platform_is_rejected(tmp_path):
    path = tmp_path / "bad_platform.xlsx"
    create_workbook(path)
    wb = load_workbook(path)
    ws = wb["Returns"]
    index = {name: i + 1 for i, name in enumerate(excel_io.COLUMNS)}
    ws.cell(row=2, column=index["order_id"]).value = "OD1"
    ws.cell(row=2, column=index["platform"]).value = "Myntra"
    wb.save(path)

    with pytest.raises(WorkbookError, match="unrecognised platform"):
        read_tasks(path)


def test_platform_names_are_case_insensitive(tmp_path):
    path = tmp_path / "case.xlsx"
    create_workbook(path)
    wb = load_workbook(path)
    ws = wb["Returns"]
    index = {name: i + 1 for i, name in enumerate(excel_io.COLUMNS)}
    ws.cell(row=2, column=index["order_id"]).value = "OD1"
    ws.cell(row=2, column=index["platform"]).value = "FLIPKART"
    wb.save(path)

    assert read_tasks(path)[0].platform == Platform.FLIPKART
